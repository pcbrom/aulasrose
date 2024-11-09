import pandas as pd
from datetime import datetime
import numpy as np
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import toml
import os

config_path = os.path.join(os.path.dirname(__file__), "config.toml")
config = toml.load(config_path)

url = config["google_sheets"]["url"]

def importar_dados(url):
    """
    Importa dados do Google Sheets a partir de uma URL fornecida.
    """
    try:
        df = pd.read_csv(url, sep='\t', dtype=str)
        print("Dados importados com sucesso.")
        return df
    except Exception as e:
        print(f"Erro ao importar dados: {e}")
        return None

def limpar_dados(df):
    df['Data'] = pd.to_datetime(df['Data'].str.strip(), format='%d/%m/%Y', errors='coerce')
    df['Nome do aluno'] = df['Nome do aluno'].str.strip()
    df['Valor da aula'] = df['Valor da aula'].str.replace(r'R\$', '', regex=True)\
                                             .str.replace(',', '.', regex=True)\
                                             .str.strip()
    df['Valor da aula'] = pd.to_numeric(df['Valor da aula'], errors='coerce')
    df = df.dropna(subset=['Data', 'Nome do aluno', 'Valor da aula'])
    return df

def limpar_dados(df):
    df['Data'] = pd.to_datetime(df['Data'].str.strip(), format='%d/%m/%Y', errors='coerce')
    df['Nome do aluno'] = df['Nome do aluno'].str.strip()
    df['Valor da aula'] = df['Valor da aula'].str.replace(r'R\$', '', regex=True)\
                                             .str.replace(',', '.', regex=True)\
                                             .str.strip()
    df['Valor da aula'] = pd.to_numeric(df['Valor da aula'], errors='coerce')
    df = df.dropna(subset=['Data', 'Nome do aluno', 'Valor da aula'])
    return df

def gerar_resumo_ano(df):
    df['mes'] = df['Data'].dt.month
    df['ano'] = df['Data'].dt.year
    resumo_ano = df.groupby(['mes', 'ano'], as_index=False)['Valor da aula'].sum()
    resumo_ano.rename(columns={'Valor da aula': 'total'}, inplace=True)
    return resumo_ano

def gerar_resumo_aluno(df):
    resumo_aluno = df.groupby('Nome do aluno')['Valor da aula'].sum().reset_index()
    resumo_aluno = resumo_aluno.sort_values(by='Valor da aula', ascending=False)
    resumo_aluno['Percentual'] = (resumo_aluno['Valor da aula'] / resumo_aluno['Valor da aula'].sum()).round(2)
    resumo_aluno['Acumulado'] = resumo_aluno['Percentual'].cumsum().round(2)
    return resumo_aluno

def gerar_saida_detalhada(df):
    linhas_saida = []
    for nome, grupo in df.groupby('Nome do aluno'):
        total = grupo['Valor da aula'].sum()
        linhas_saida.append({'Data': f"Nome: {nome}", 'Valor': ''})
        for _, row in grupo.iterrows():
            data_formatada = row['Data'].strftime('%d/%m/%Y')
            valor_formatado = f"R$ {row['Valor da aula']:.2f}"
            linhas_saida.append({'Data': data_formatada, 'Valor': valor_formatado})
        linhas_saida.append({'Data': 'Total', 'Valor': f"R$ {total:.2f}"})
    return pd.DataFrame(linhas_saida)

def salvar_em_excel(saida, filename=None):
    """
    Salva os dados em um arquivo Excel com formatação.
    """
    if not filename:
        filename = f'relatorio_{datetime.now().date()}.xlsx'
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            saida.to_excel(writer, index=False, header=False, sheet_name='Detalhes')
            workbook = writer.book
            worksheet = workbook['Detalhes']

            # Ajustar largura das colunas
            worksheet.column_dimensions['A'].width = 20
            worksheet.column_dimensions['B'].width = 20

            alignment = Alignment(horizontal='center')
            bold_font = Font(bold=True)
            thick_border = Border(left=Side(style='thick'), right=Side(style='thick'),
                                  top=Side(style='thick'), bottom=Side(style='thick'))
            light_gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

            for row_idx, row in enumerate(worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=2), start=1):
                for cell in row:
                    cell.alignment = alignment
                    if str(cell.value).startswith("Nome:"):
                        cell.font = bold_font
                        cell.fill = light_gray_fill
                        worksheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
                    if str(cell.value).startswith("Total"):
                        cell.font = bold_font

        print(f"Relatório salvo com sucesso em {filename}")
    except Exception as e:
        print(f"Erro ao salvar o relatório: {e}")

def processar_dados(data_inicio, data_fim):
    """
    Função principal para processar os dados.
    """
    df = importar_dados(url)
    if df is None:
        return None, None, None

    df = limpar_dados(df)

    # Converter data_inicio e data_fim para datetime64
    data_inicio = pd.to_datetime(data_inicio)
    data_fim = pd.to_datetime(data_fim)

    # Filtrar o DataFrame pelas datas
    df = df[(df['Data'] >= data_inicio) & (df['Data'] <= data_fim)]

    resumo_ano = gerar_resumo_ano(df)
    resumo_aluno = gerar_resumo_aluno(df)
    saida = gerar_saida_detalhada(df)

    return saida, resumo_ano, resumo_aluno
